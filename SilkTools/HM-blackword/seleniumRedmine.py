#!/usr/bin/python3
# -*-coding:utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.by import By
import re, string
import config as CF
import xlwt
from selenium.webdriver.chrome.options import Options
import xlrd
from openpyxl import load_workbook
import time
import xlsxwriter
from xlutils.copy import copy
from sendReport import SENDMAIL

import xlutils, openpyxl


class SeleniumRedmine(object):

    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')

        self.driver = webdriver.Chrome(chrome_options=chrome_options)
        self.driver.get(CF.base_url)

        self.driver.find_element_by_id('username').send_keys('alexander.zhao')
        self.driver.find_element_by_id('password').send_keys('Mi123456')
        self.driver.find_element_by_xpath('//input[@type="submit"]').click()

    def getEachNumber(self):
        for key in CF.UATData:
            self.driver.get(CF.UATData[key])
            number = self.driver.find_element_by_xpath('//p[@class="pagination"]/span[@class="items"]').text
            num = self.getNumber(number)
            CF.UATList[key] = num
            print(key, num)

    def getAppEachNumber(self):
        for key in CF.AppData:
            self.driver.get(CF.AppData[key])
            try:
                number = self.driver.find_element_by_xpath('//p[@class="pagination"]/span[@class="items"]').text
                num = self.getNumber(number)
                CF.APPList[key] = num
                print(key, num)
            except:
                pass

    def getNumber(self, content):
        number = content.split('/')[1]
        number = number.replace(')', '')
        return number

    def quit(self):
        self.driver.quit()

    def writeToExcel(self):
        file = CF.filepath + r'HM-BUG统计-.xlsx'
        _date = time.strftime('%Y%m%d', time.localtime(time.time()))
        print(_date)
        newfile = 'HM-BUG统计-' + _date + '.xlsx'
        print(newfile)

        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # 写入UAT数据
        sheet.cell(2, 2, int(CF.UATList['Low']))
        sheet.cell(3, 2, int(CF.UATList['Normal']))
        sheet.cell(4, 2, int(CF.UATList['High']))
        sheet.cell(5, 2, int(CF.UATList['Urgent']))
        sheet.cell(6, 2, int(CF.UATList['Immediate']))

        # 写入App数据

        sheet.cell(25, 2, int(CF.APPList['Low']))
        sheet.cell(26, 2, int(CF.APPList['Normal']))
        sheet.cell(27, 2, int(CF.APPList['High']))
        sheet.cell(28, 2, int(CF.APPList['Urgent']))
        sheet.cell(29, 2, int(CF.APPList['Immediate']))

        newfile = CF.filepath + newfile
        workbook.save(newfile)

        SENDMAIL(CF.sendTo,CF.sendCc, newfile)


if __name__ == '__main__':
    redmine = SeleniumRedmine()
    redmine.getEachNumber()
    print('***********************************')
    redmine.getAppEachNumber()
    print(CF.UATList)
    print(CF.APPList)
    redmine.quit()

    redmine.writeToExcel()
